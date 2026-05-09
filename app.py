from flask import Flask, render_template, request, redirect, session, send_from_directory, Response
from flask_sqlalchemy import SQLAlchemy
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
import os

app = Flask(__name__)
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "uploads")
app.secret_key = os.environ.get("SECRET_KEY", "dev-secret-key")

DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///classroom.db")
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# ================= MODELS =================

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100))
    password = db.Column(db.String(100))
    role = db.Column(db.String(20))
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=True)

class Classroom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100))

class Assignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(100))
    classroom_id = db.Column(db.Integer)

class Submission(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student = db.Column(db.String(100))
    assignment = db.Column(db.String(100))
    content = db.Column(db.String(300))
    file = db.Column(db.String(300))
    score = db.Column(db.Integer, default=0)

class Notification(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student = db.Column(db.String(100))
    message = db.Column(db.String(300))
    is_read = db.Column(db.Boolean, default=False)

# ใหม่: เนื้อหา/ลิงก์ที่ครูโพสต์
class Post(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=False)
    title = db.Column(db.String(200))
    content = db.Column(db.String(1000))
    link = db.Column(db.String(500), nullable=True)

# ================= LOGIN =================

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        user = User.query.filter_by(
            username=request.form["username"],
            password=request.form["password"]
        ).first()
        if user:
            session["username"] = user.username
            session["role"] = user.role
            if user.role == "teacher":
                return redirect("/teacher")
            return redirect("/student")
    return render_template("login.html")

# ================= REGISTER =================

@app.route("/register", methods=["GET", "POST"])
def register():
    if request.method == "POST":
        new_user = User(
            username=request.form["username"],
            password=request.form["password"],
            role=request.form["role"],
            classroom_id=request.form.get("classroom_id") or None
        )
        db.session.add(new_user)
        db.session.commit()
        return redirect("/")
    classrooms = Classroom.query.all()
    return render_template("register.html", classrooms=classrooms)

# ================= LOGOUT =================

@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")

# ================= TEACHER =================

@app.route("/teacher")
def teacher():
    classrooms = Classroom.query.all()
    assignments = Assignment.query.all()
    submissions = Submission.query.all()
    posts = Post.query.all()

    # นับและดึงชื่อนักเรียนแต่ละห้อง
    classroom_students = {}
    for c in classrooms:
        students = User.query.filter_by(role="student", classroom_id=c.id).all()
        classroom_students[c.id] = students

    # นับคนที่ยังไม่ส่งแต่ละงาน
    missing_counts = {}
    for assignment in assignments:
        students_in_room = User.query.filter_by(
            role="student", classroom_id=assignment.classroom_id
        ).all()
        submitted_students = {
            s.student for s in Submission.query.filter_by(assignment=assignment.title).all()
        }
        missing = [s for s in students_in_room if s.username not in submitted_students]
        missing_counts[assignment.id] = len(missing)

    return render_template(
        "teacher.html",
        classrooms=classrooms,
        assignments=assignments,
        submissions=submissions,
        missing_counts=missing_counts,
        classroom_students=classroom_students,
        posts=posts
    )

@app.route("/create_classroom", methods=["POST"])
def create_classroom():
    room = Classroom(name=request.form["name"])
    db.session.add(room)
    db.session.commit()
    return redirect("/teacher")

# ใหม่: ลบห้องเรียน
@app.route("/delete_classroom/<int:id>")
def delete_classroom(id):
    classroom = db.session.get(Classroom, id)
    if classroom:
        # ย้ายนักเรียนในห้องออกก่อน
        User.query.filter_by(classroom_id=id).update({"classroom_id": None})
        # ลบ assignments และ posts ของห้องนี้
        Assignment.query.filter_by(classroom_id=id).delete()
        Post.query.filter_by(classroom_id=id).delete()
        db.session.delete(classroom)
        db.session.commit()
    return redirect("/teacher")

@app.route("/create_assignment", methods=["POST"])
def create_assignment():
    work = Assignment(
        title=request.form["title"],
        classroom_id=request.form["classroom_id"]
    )
    db.session.add(work)
    db.session.commit()
    return redirect("/teacher")

# ลบงาน (Assignment)
@app.route("/delete_assignment/<int:id>")
def delete_assignment(id):
    assignment = db.session.get(Assignment, id)
    if assignment:
        db.session.delete(assignment)
        db.session.commit()
    return redirect("/teacher")

# ลบงานที่นักเรียนส่ง (Submission)
@app.route("/delete_submission/<int:id>")
def delete_submission(id):
    submission = db.session.get(Submission, id)
    if submission:
        if submission.file:
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], submission.file)
            if os.path.exists(file_path):
                os.remove(file_path)
        db.session.delete(submission)
        db.session.commit()
    return redirect("/teacher")

# ใหม่: โพสต์เนื้อหา/ลิงก์
@app.route("/create_post", methods=["POST"])
def create_post():
    post = Post(
        classroom_id=request.form["classroom_id"],
        title=request.form["title"],
        content=request.form.get("content", ""),
        link=request.form.get("link") or None
    )
    db.session.add(post)
    db.session.commit()
    return redirect("/teacher")

# ใหม่: ลบโพสต์
@app.route("/delete_post/<int:id>")
def delete_post(id):
    post = db.session.get(Post, id)
    if post:
        db.session.delete(post)
        db.session.commit()
    return redirect("/teacher")

# ================= แจ้งเตือนคนที่ยังไม่ส่งงาน =================

@app.route("/notify_missing/<int:assignment_id>")
def notify_missing(assignment_id):
    assignment = db.session.get(Assignment, assignment_id)
    if not assignment:
        return redirect("/teacher")

    students_in_room = User.query.filter_by(
        role="student", classroom_id=assignment.classroom_id
    ).all()
    submitted_students = {
        s.student for s in Submission.query.filter_by(assignment=assignment.title).all()
    }
    for student in students_in_room:
        if student.username not in submitted_students:
            db.session.add(Notification(
                student=student.username,
                message=f"คุณยังไม่ได้ส่งงาน: {assignment.title} กรุณาส่งก่อนกำหนด!"
            ))
    db.session.commit()
    return redirect("/teacher")

# ================= EXPORT SCORES =================

@app.route("/export_scores")
def export_scores():
    classrooms = Classroom.query.all()
    assignments = Assignment.query.all()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center")

    for classroom in classrooms:
        ws = wb.create_sheet(title=classroom.name[:31])
        room_assignments = [a for a in assignments if a.classroom_id == classroom.id]
        headers = ["นักเรียน"] + [a.title for a in room_assignments] + ["รวม", "เฉลี่ย"]

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        submissions = Submission.query.filter(
            Submission.assignment.in_([a.title for a in room_assignments])
        ).all() if room_assignments else []

        score_map = {}
        for sub in submissions:
            if sub.student not in score_map:
                score_map[sub.student] = {}
            score_map[sub.student][sub.assignment] = sub.score

        for row_idx, (student, scores) in enumerate(score_map.items(), 2):
            ws.cell(row=row_idx, column=1, value=student)
            total = 0; count = 0
            for col_idx, assignment in enumerate(room_assignments, 2):
                score = scores.get(assignment.title, "-")
                ws.cell(row=row_idx, column=col_idx, value=score)
                if isinstance(score, int):
                    total += score; count += 1
            ws.cell(row=row_idx, column=len(headers) - 1, value=total)
            ws.cell(row=row_idx, column=len(headers), value=round(total/count, 2) if count > 0 else "-")

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=scores.xlsx"}
    )

# ================= STUDENT =================

@app.route("/student")
def student():
    assignments = Assignment.query.all()
    submissions = Submission.query.filter_by(student=session["username"]).all()
    notifications = Notification.query.filter_by(
        student=session["username"], is_read=False
    ).all()

    # ดึงโพสต์ของห้องนักเรียน
    current_user = User.query.filter_by(username=session["username"]).first()
    posts = []
    classroom_name = None
    if current_user and current_user.classroom_id:
        posts = Post.query.filter_by(classroom_id=current_user.classroom_id).all()
        room = db.session.get(Classroom, current_user.classroom_id)
        classroom_name = room.name if room else None

    return render_template(
        "student.html",
        assignments=assignments,
        submissions=submissions,
        notifications=notifications,
        posts=posts,
        current_user=current_user,
        classroom_name=classroom_name
    )

@app.route("/read_notification/<int:id>")
def read_notification(id):
    notif = db.session.get(Notification, id)
    if notif and notif.student == session["username"]:
        notif.is_read = True
        db.session.commit()
    return redirect("/student")

# ใหม่: ออกจากห้องเรียน
@app.route("/leave_classroom")
def leave_classroom():
    user = User.query.filter_by(username=session["username"]).first()
    if user:
        user.classroom_id = None
        db.session.commit()
    return redirect("/student")

# นักเรียนลบงานของตัวเอง
@app.route("/student_delete/<int:id>")
def student_delete(id):
    submission = db.session.get(Submission, id)
    if submission and submission.student == session["username"]:
        if submission.file:
            file_path = os.path.join(app.config["UPLOAD_FOLDER"], submission.file)
            if os.path.exists(file_path):
                os.remove(file_path)
        db.session.delete(submission)
        db.session.commit()
    return redirect("/student")
@app.route("/submit", methods=["POST"])
def submit():
    uploaded_file = request.files["file"]
    filename = None
    if uploaded_file and uploaded_file.filename != "":
        filename = uploaded_file.filename
        os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
        uploaded_file.save(os.path.join(app.config["UPLOAD_FOLDER"], filename))

    new_submit = Submission(
        student=session["username"],
        assignment=request.form["assignment"],
        content=request.form["content"],
        file=filename
    )
    db.session.add(new_submit)
    db.session.commit()
    return redirect("/student")

# ================= เปิดไฟล์ =================

@app.route("/uploads/<filename>")
def uploaded_file(filename):
    ext = filename.rsplit(".", 1)[-1].lower()
    inline_types = {
        "pdf": "application/pdf", "png": "image/png",
        "jpg": "image/jpeg", "jpeg": "image/jpeg",
        "gif": "image/gif", "webp": "image/webp", "txt": "text/plain",
    }
    if ext in inline_types:
        return send_from_directory(app.config["UPLOAD_FOLDER"], filename,
                                   mimetype=inline_types[ext], as_attachment=False)
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename, as_attachment=True)

# ================= GRADE =================

@app.route("/grade/<int:id>", methods=["POST"])
def grade(id):
    submission = db.session.get(Submission, id)
    submission.score = request.form["score"]
    db.session.commit()
    return redirect("/teacher")

# ================= MAIN =================

if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(host="0.0.0.0", port=5000, debug=True)
